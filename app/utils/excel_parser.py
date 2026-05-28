"""
Excel file parsing utilities
"""

import logging
from typing import Any

from openpyxl import load_workbook


logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Parse Excel files for student bulk uploads
    """

    REQUIRED_COLUMNS = {"name", "email", "batch", "password", "category"}

    @staticmethod
    def parse_file(file_path: str) -> tuple[list[dict[str, Any]], list[str]]:
        """
        Parse Excel file and extract student data

        Args:
            file_path: Path to Excel file

        Returns:
            Tuple of (data_list, errors)
            - data_list: List of dictionaries with student data
            - errors: List of error messages
        """
        errors = []
        data = []

        try:
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            if not worksheet:
                errors.append("No active worksheet found in Excel file")
                return [], errors

            # Get headers
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(cell.value.lower().strip())

            # Validate headers
            header_set = set(headers)
            missing_columns = ExcelParser.REQUIRED_COLUMNS - header_set
            if missing_columns:
                errors.append(
                    f"Missing required columns: {', '.join(missing_columns)}"
                )
                return [], errors

            # Parse rows
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                row_data = {}
                try:
                    for col_idx, header in enumerate(headers):
                        row_data[header] = row[col_idx] if col_idx < len(row) else None

                    # Validate required fields
                    for required_field in ExcelParser.REQUIRED_COLUMNS:
                        if not row_data.get(required_field):
                            errors.append(
                                f"Row {row_idx}: Missing required field '{required_field}'"
                            )
                            continue

                    if row_data.get("name"):
                        data.append(row_data)

                except Exception as e:
                    errors.append(f"Row {row_idx}: Error parsing row - {str(e)}")
                    logger.error(f"Error parsing row {row_idx}: {str(e)}")

            workbook.close()

        except FileNotFoundError:
            errors.append(f"File not found: {file_path}")
        except Exception as e:
            errors.append(f"Error parsing Excel file: {str(e)}")
            logger.error(f"Error parsing Excel file: {str(e)}")

        return data, errors
